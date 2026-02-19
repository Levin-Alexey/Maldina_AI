#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Импорт инструкций по устранению неисправностей из files/inst.xlsx в БД
Запуск: python import_instructions.py
"""

import asyncio
import hashlib
import json
from pathlib import Path
import openpyxl
from sentence_transformers import SentenceTransformer
from sqlalchemy import text

from db import SessionLocal
from models import TroubleshootInstruction

# Модель для эмбеддингов (та же, что в kb_search и product_search)
model = SentenceTransformer("paraphrase-multilingual-MiniLM-L12-v2")

# Путь к файлу с инструкциями
EXCEL_FILE = Path("files/inst.xlsx")

# Пороговое расстояние между стоками для парсинга (иначе считаем за одно значение)
MAX_ROWS_PER_CELL = 5


def parse_multiline_text(cell_value):
    """
    Парсит текст из ячейки Excel, где значения разделены переносами строк.
    Пример: "а_МР9_синий\nа_МР9_синий_v2\nа_МР9_красный"
    Возвращает строку через запятую: "а_МР9_синий, а_МР9_синий_v2, а_МР9_красный"
    """
    if not cell_value:
        return None
    
    # Если это уже не строка - преобразуем
    cell_value = str(cell_value).strip()
    if not cell_value:
        return None
    
    # Разбиваем по переносам строк и очищаем
    items = [item.strip() for item in cell_value.split('\n') if item.strip()]
    
    return ', '.join(items) if items else None


def collect_steps(row_data):
    """
    Собирает шаги из ячеек в JSON формат.
    row_data - список значений из строки Excel
    Ожидаемая структура: [..., шаг_1_индекс, шаг_2_индекс, шаг_3_индекс, ...]
    """
    steps = {}
    
    # Колонки F-M это шаги 1-8 (индексы 5-12 в 0-базированном массиве)
    # Row = [SKU_внутренний, SKU_WB, SKU_Ozon, Название, Проблема, Шаг1, Шаг2, ...]
    for i, step_value in enumerate(row_data[5:13], start=1):  # Индексы 5-12 = колонки F-M
        if step_value and str(step_value).strip():
            steps[str(i)] = str(step_value).strip()
    
    return steps


def create_content_hash(product_name, issue_description):
    """
    Создает хеш для дедупликации.
    Хеш основан на product_name + issue_description
    """
    hash_text = f"{product_name}|{issue_description}".lower()
    hash_obj = hashlib.sha256(hash_text.encode('utf-8'))
    return hash_obj.hexdigest()[:32]  # Берем первые 32 символа


def create_embedding(product_name, issue_description):
    """
    Создает эмбеддинг для семантического поиска.
    Комбинирует имя товара и описание проблемы.
    """
    emb_text = f"{product_name}: {issue_description}"
    embedding = model.encode(emb_text)
    return embedding.tolist()


async def import_from_excel():
    """Основная функция импорта из Excel"""
    
    print("=" * 70)
    print("  ИМПОРТ ИНСТРУКЦИЙ ПО УСТРАНЕНИЮ НЕИСПРАВНОСТЕЙ")
    print("=" * 70)
    print()
    
    # Проверяем наличие файла
    if not EXCEL_FILE.exists():
        print(f"❌ Файл не найден: {EXCEL_FILE}")
        return
    
    print(f"📂 Открываю файл: {EXCEL_FILE}")
    
    # Открываем Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    print(f"📋 Лист: {ws.title}")
    print(f"📊 Размер: {ws.max_row} строк × {ws.max_column} колонок")
    print()
    
    # Подключаемся к БД
    async with SessionLocal() as session:
        imported = 0
        skipped = 0
        errors = 0
        
        # Пропускаем заголовок (строка 1) и пустые строки
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                # Получаем значения ячеек
                row_values = [cell.value for cell in row]
                
                # Проверяем, не пустая ли строка
                if not any(row_values):
                    continue
                
                # Парсим артикулы
                internal_sku = parse_multiline_text(row_values[0])
                wb_sku = parse_multiline_text(row_values[1])
                ozon_sku = parse_multiline_text(row_values[2])
                
                product_name = str(row_values[3]).strip() if row_values[3] else None
                issue_description = str(row_values[4]).strip() if row_values[4] else None
                
                # Пропускаем если нет необходимых данных
                if not product_name or not issue_description:
                    skipped += 1
                    continue
                
                # Собираем шаги
                steps = collect_steps(row_values)
                
                if not steps:
                    print(f"⚠️  Строка {row_idx}: Нет шагов, пропускаю")
                    skipped += 1
                    continue
                
                # Создаем хеш и эмбеддинг
                content_hash = create_content_hash(product_name, issue_description)
                embedding = create_embedding(product_name, issue_description)
                
                # Проверяем дубликат по хешу
                existing = await session.execute(
                    text("SELECT id FROM troubleshoot_instructions WHERE content_hash = :hash"),
                    {"hash": content_hash}
                )
                if existing.scalar():
                    print(f"⏭️  Строка {row_idx}: Дубликат (уже в БД), пропускаю")
                    skipped += 1
                    continue
                
                # Создаем запись
                instruction = TroubleshootInstruction(
                    internal_sku=internal_sku,
                    wb_sku=wb_sku,
                    ozon_sku=ozon_sku,
                    product_name=product_name,
                    issue_description=issue_description,
                    steps=steps,
                    content_hash=content_hash,
                    embedding=embedding
                )
                
                session.add(instruction)
                await session.flush()  # Получаем ID
                
                print(f"✅ Строка {row_idx}: {product_name[:40]:40} | {issue_description[:30]:30}")
                imported += 1
                
                # Коммитим каждые 10 записей
                if imported % 10 == 0:
                    await session.commit()
                    print(f"   💾 Промежуточное сохранение ({imported} записей)")
                
            except Exception as e:
                errors += 1
                print(f"❌ Ошибка в строке {row_idx}: {e}")
                await session.rollback()
                continue
        
        # Финальный коммит
        if imported > 0:
            await session.commit()
            print()
            print("=" * 70)
            print(f"✅ ИМПОРТ ЗАВЕРШЕН")
            print(f"   ✨ Импортировано: {imported}")
            print(f"   ⏭️  Пропущено: {skipped}")
            print(f"   ❌ Ошибок: {errors}")
            print(f"   📊 Всего: {imported + skipped} строк обработано")
            print("=" * 70)
        else:
            print("❌ Ничего не импортировано!")
            await session.rollback()


async def check_imported_data():
    """Проверяет импортированные данные"""
    async with SessionLocal() as session:
        result = await session.execute(
            text("SELECT COUNT(*) FROM troubleshoot_instructions")
        )
        count = result.scalar()
        
        if count > 0:
            print()
            print("📊 Статистика БД:")
            print(f"   Всего инструкций: {count}")
            
            # Выборка нескольких записей для проверки
            result = await session.execute(
                text("""
                    SELECT id, product_name, issue_description, 
                           array_length(akeys(steps), 1) as steps_count
                    FROM troubleshoot_instructions 
                    LIMIT 3
                """)
            )
            rows = result.fetchall()
            
            print()
            print("   Примеры записей:")
            for row in rows:
                print(f"     ID {row[0]}: {row[1]} → {row[2][:40]} ({row[3]} шагов)")


async def main():
    """Главная функция"""
    await import_from_excel()
    await check_imported_data()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n⚠️  Импорт прерван пользователем")
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
