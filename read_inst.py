import openpyxl

wb = openpyxl.load_workbook('files/inst.xlsx')
ws = wb.active

print(f"Всего листов: {len(wb.sheetnames)}")
print(f"Активный лист: {ws.title}")
print(f"Размер: {ws.max_row} строк, {ws.max_column} колонок")
print("\n" + "="*80 + "\n")

for row in ws.iter_rows(values_only=True):
    print(row)
