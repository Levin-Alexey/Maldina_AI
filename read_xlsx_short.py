#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl

wb = openpyxl.load_workbook('files/inst.xlsx')

print(f"Листы в файле: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*100}")
    print(f"ЛИСТ: {sheet_name}")
    print(f"Размер: {ws.max_row} строк × {ws.max_column} колонок")
    print(f"{'='*100}\n")
    
    # Показать только первые 15 строк
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 15:
            print(f"\n... (всего строк: {ws.max_row})\n")
            break
        row_text = ' | '.join([str(cell) if cell is not None else '' for cell in row])
        print(f"{i:3d}: {row_text}")
