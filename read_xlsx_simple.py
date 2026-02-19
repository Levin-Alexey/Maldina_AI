#!/usr/bin/env python
# -*- coding: utf-8 -*-

try:
    import openpyxl
    print("✓ openpyxl доступен")
except:
    print("✗ openpyxl НЕ установлен")
    import sys
    sys.exit(1)

wb = openpyxl.load_workbook('files/inst.xlsx')

print(f"\nЛисты в файле: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"ЛИСТ: {sheet_name}")
    print(f"Размер: {ws.max_row} строк × {ws.max_column} колонок")
    print(f"{'='*80}\n")
    
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 50:  # Ограничим вывод 50 строками
            print(f"... (ещё {ws.max_row - 50} строк)")
            break
        print(f"Строка {i}: {row}")
